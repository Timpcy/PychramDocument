#可参考的教程
##https://openpyxl.readthedocs.io/en/stable/styles.html  官方文件
##https://blog.csdn.net/sinat_28576553/article/details/81275650

import openpyxl
workbook=openpyxl.load_workbook("/Users/Tim/Desktop/浙江美元债情况.xlsx")
shenames=workbook.sheetnames
print(shenames)
#还可以通过如下写法获得表对象
ws=workbook[shenames[0]]
print(ws)
ws1=workbook[shenames[1]]
print(ws1)
ws2=workbook[shenames[2]]
print(ws2)
print(ws.max_row)
print(ws1.max_row)
print(ws.cell(row=3,column=3).value)
print(ws1.cell(row=1,column=1).value)

#copy title
for i in range(1,ws1.max_column):
     ws2.cell(row=1,column=i).value=ws1.cell(row=1,column=i).value

foundedvalue=[]
n=2#是一个计数器 从2开始 第一行为标题
for i in range(1,ws.max_row):#i 表1所有行数的列举
    for j in range(1,ws1.max_row):#j 表1所有行数的列举
        #此处搜索的范围被锁定在第一列；column=1；找到相同的值
        if ws.cell(row=i,column=3).value==ws1.cell(row=j,column=16).value or ws.cell(row=i,column=2).value==ws1.cell(row=j,column=19).value:
            #输出被发现的值并进行储存
            #并且需要排除None 所在单元格造成的印象
            if ws.cell(row=i,column=3).value==None:
                pass
            else:
                print(ws.cell(row=i,column=3).value)
                if ws.cell(row=i,column=3).value in foundedvalue:
                    pass
                else:
                    foundedvalue.append(ws.cell(row=i,column=3).value)
                    #负值函数
                    for col in range(1,ws1.max_column):
                        #在表4中进行副职，以表2中的值为准
                        ws2.cell(row=n,column=col).value=ws1.cell(row=j,column=col).value
                    n=n+1
        else:
            continue
workbook.save (filename='浙江美元债情况.xlsx')